import openpyxl
import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
import time

# Load the Google Cloud credentials from a service account key file
credentials = service_account.Credentials.from_service_account_file('evident-torus-378319-06bfd86395e7.json')

# Set up the Google Custom Search API client
api_key = 'AIzaSyClT0ofnUr8fOkIgCoPTpHWmjvmfkVse24'
search_engine_id = '8131cc44845744a14'
service = build('customsearch', 'v1', developerKey=api_key)

# Load the Excel workbook and select the first worksheet
workbook = openpyxl.load_workbook('netflix.xlsx')
sheet = workbook.active

# Add a header for the image URL column
sheet.cell(row=1, column=sheet.max_column+1).value = 'imageurl'

# Iterate over the rows in the worksheet and search for the game images
index = 1
for row in sheet.iter_rows(min_row=2):
    if index % 100 == 0 and index > 0:
        print("Pausing for 1 minute to avoid rate limit.")
        time.sleep(80)  # Pause for 1 minute
    movie_name = row[2].value
    movie_type = row[1].value
    movie_year =  str(row[7].value)
    query = movie_name + ' ' + movie_type + ' ' + movie_year + ' poster'
    res = service.cse().list(q=query, cx=search_engine_id, searchType='image').execute()
    print(f'Added image for {movie_name}' )
    # Save the URL of the first image in the results to the Excel sheet
    if 'items' in res and len(res['items']) > 0:
        for item in res['items']:
            if item['link'].startswith('https://'):
                image_url = item['link']
                print(image_url)
                sheet.cell(row=row[0].row, column=sheet.max_column).value = image_url
                break
    index += 1

# Save the modified Excel workbook
workbook.save('netflix_with_images.xlsx')